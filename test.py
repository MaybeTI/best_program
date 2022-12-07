from openpyxl import load_workbook

fn = 'result.xlsx'

wb = load_workbook(fn)

data = "04.12.2022"


ws = wb.create_sheet(data)
ws.column_dimensions["A"].width = 30




ws["A5"] = "lol"

wb.save(fn)
wb.close()