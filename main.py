import requests
import json
from openpyxl import load_workbook
from datetime import datetime

days = 1

fn1 = 'Завтра.xlsx'
fn2 = "Послезавтра.xlsx"
fn3 = "result.xlsx"

if days == 1:
    wb = load_workbook(fn1)
elif days == 2:
    wb = load_workbook(fn2)
else:
    wb = load_workbook(fn3)

headers = {"x-fsign": "SW9D1eZo"}

result = []

def visc(feed, name1, name2, liga = 1):
    f = f"df_to_5_{feed}_1"
    url = f"https://d.flashscore.ua/x/feed/{f}"
    response = requests.get(url=url, headers=headers)
    data = response.text.split("¬")
    data_list1 = [{}]


    for i in data:
        key = i.split("÷")[0]
        value = i.split("÷")[-1]

        if "~" in key:
            data_list1.append({key: value})
        else:
            data_list1[-1].update({key: value})

    for element in data_list1:
        if "TR" in list(element.keys())[0] and bet(feed):

            if int(element["~TR"]) < 3 and element["TN"] == name1 and int(element["TM"]) * 2 <= int(element["TG"][:element["TG"].index(":")]):
                if name1 not in result:
                    result.append({name1: liga})

            elif int(element["~TR"]) < 3 and element["TN"] == name2 and int(element["TM"]) * 2 <= int(element["TG"][:element["TG"].index(":")]):
                if name2 not in result:
                    result.append({name2: liga})

def bet(feed):
    f = f"df_dos_5_{feed}_"
    url = f"https://d.flashscore.ua/x/feed/{f}"
    response = requests.get(url=url, headers=headers)
    data = response.text.split("¬")
    if data[3].count("false") == 3:
        return False
    else:
        return True

def main():
    feed = f'f_1_{days}_2_ua_5'
    url = f"https://d.flashscore.ua/x/feed/{feed}"
    response = requests.get(url=url, headers=headers)
    data = response.text.split("¬")

    data_list = [{}]

    for item in data:
        key = item.split("÷")[0]
        value = item.split("÷")[-1]

        if "~AA" in key:
            data_list.append({key: value})
        else:
            data_list[-1].update({key: value})

    for item in data_list:
        if "AE" in item:
            name1 = item["AE"]
            if name1 == "Колвін-Бей":
                print(item)
        if "AF" in item:
            name2 = item["AF"]
        if "~ZA" in item:
            liga = item.get("~ZA")

        if "~AA" in item and name1[-1] != "Ж" and name2[-1] != "Ж" and "U" not in name1 and "U" not in name2 and\
                "U"not in liga and "Жінки" not in name1 and "жінки" not in name1 and "Жінки" not in name2 and\
                "жінки" not in name2 and "Жінки" not in liga and "жінки" not in liga:

            visc(item["~AA"], name1, name2, liga)


    n = 0

    day = datetime.fromtimestamp(int(data_list[1]["AD"]))

    a = str(day)

    ws = wb.create_sheet(a[:10])

    for j in wb.sheetnames[:1]:
        del wb[j]

    for i in result:
        for key in i.keys():
            if key[-1] != "Ж" and "U" not in key and "U" not in i[key]:

                ws.column_dimensions["A"].width = 25
                ws.column_dimensions["B"].width = 60
                n += 1
                ws[f"A{n}"] = key
                ws[f"B{n}"] = i[key]

if __name__ == '__main__':
    main()

if days == 1:
    wb.save(fn1)
    wb.close()
elif days == 2:
    wb.save(fn2)
    wb.close()
else:
    wb.save(fn3)
    wb.close()
